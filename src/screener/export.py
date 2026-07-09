import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from src.screener.engine import ScreenerEngine

# --------------------------------------------------
# Create Output Folder
# --------------------------------------------------

os.makedirs("output", exist_ok=True)

# --------------------------------------------------
# Load Engine
# --------------------------------------------------

engine = ScreenerEngine()

wb = Workbook()

# Remove default sheet
wb.remove(wb.active)

green = PatternFill(
    start_color="90EE90",
    end_color="90EE90",
    fill_type="solid"
)

red = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

# --------------------------------------------------
# Export Every Preset
# --------------------------------------------------

for preset in engine.presets():

    df = engine.apply_filters(preset)

    ws = wb.create_sheet(title=preset[:31])

    # Write Header
    for col, name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col).value = name

    # Write Data
    for r, row in enumerate(df.itertuples(index=False), start=2):

        for c, value in enumerate(row, start=1):

            cell = ws.cell(row=r, column=c)

            cell.value = value

    # --------------------------------------------------
    # Colour Composite Score
    # --------------------------------------------------

    if "composite_quality_score" in df.columns:

        score_col = list(df.columns).index(
            "composite_quality_score"
        ) + 1

        for row in range(2, ws.max_row + 1):

            value = ws.cell(row, score_col).value

            if value is None:
                continue

            if value >= 75:
                ws.cell(row, score_col).fill = green

            else:
                ws.cell(row, score_col).fill = red

print("Saving Excel...")

wb.save("output/screener_output.xlsx")

print("Done.")

print("Sheets Created:")

for s in wb.sheetnames:
    print("-", s)