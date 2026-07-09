import os
import sqlite3
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill

print("=" * 70)
print("DAY 20 - PEER COMPARISON REPORT")
print("=" * 70)

os.makedirs("output", exist_ok=True)

conn = sqlite3.connect("db/nifty100.db")

peer = pd.read_sql(
    "SELECT * FROM peer_percentiles",
    conn
)

company = pd.read_sql(
    """
    SELECT
        id AS company_id,
        company_name
    FROM companies
    """,
    conn
).drop_duplicates()

wb = Workbook()
wb.remove(wb.active)

green = PatternFill(
    start_color="90EE90",
    end_color="90EE90",
    fill_type="solid"
)

yellow = PatternFill(
    start_color="FFF59D",
    end_color="FFF59D",
    fill_type="solid"
)

red = PatternFill(
    start_color="FFCDD2",
    end_color="FFCDD2",
    fill_type="solid"
)

groups = sorted(peer["peer_group_name"].dropna().unique())

for group in groups:

    ws = wb.create_sheet(group[:31])

    group_df = peer[
        peer["peer_group_name"] == group
    ].copy()

    table = group_df.pivot_table(
        index=["company_id", "year"],
        columns="metric",
        values="percentile_rank",
        aggfunc="first"
    ).reset_index()

    table = table.merge(
        company,
        on="company_id",
        how="left"
    )

    cols = ["company_id", "company_name", "year"] + [
        c for c in table.columns
        if c not in ["company_id", "company_name", "year"]
    ]

    table = table[cols]

    for c, col in enumerate(table.columns, start=1):
        ws.cell(row=1, column=c).value = col

    for r, row in enumerate(table.itertuples(index=False), start=2):

        for c, value in enumerate(row, start=1):

            cell = ws.cell(row=r, column=c)

            cell.value = value

            if isinstance(value, (int, float)):

                if value >= 0.75:
                    cell.fill = green

                elif value >= 0.25:
                    cell.fill = yellow

                else:
                    cell.fill = red

print()

print("Sheets Created:", len(wb.sheetnames))

wb.save("output/peer_comparison.xlsx")

print("Excel Saved Successfully.")

conn.close()